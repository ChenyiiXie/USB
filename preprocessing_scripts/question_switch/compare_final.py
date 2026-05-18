#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
question - 
question（q1.1, q1.2）sheet，
"""

import json
import os
import pandas as pd
from pathlib import Path
import argparse
from typing import Dict, List, Any

# （Project root）
DEFAULT_ORIGINAL_DIR = "nuscenes_variants_by_shard_impromptu_full"
DEFAULT_CONVERTED_DIR = "nuscenes_switch"
DEFAULT_OUTPUT_FILE = "comparison_results_final.xlsx"

def load_json_files(directory: str) -> List[Dict]:
    """JSON"""
    all_data = []
    json_files = []
    
    # JSON
    for file_path in Path(directory).glob("*.json"):
        json_files.append(file_path)
    
    print(f"Found {len(json_files)} JSON files")
    
    # 
    json_files.sort()
    
    for file_path in json_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    all_data.extend(data)
                else:
                    all_data.append(data)
            print(f": {file_path.name}")
        except Exception as e:
            print(f"failed {file_path.name}: {e}")
    
    return all_data

def get_main_question_type(question_type: str) -> str:
    """question，"""
    if question_type.startswith('q1'):
        return 'q1'
    elif question_type.startswith('q2'):
        return 'q2'
    elif question_type.startswith('q3'):
        return 'q3'
    elif question_type.startswith('q4'):
        return 'q4'
    elif question_type.startswith('q5'):
        return 'q5'
    elif question_type.startswith('q6'):
        return 'q6'
    elif question_type.startswith('q7'):
        return 'q7'
    else:
        return question_type

def create_final_comparison_data(original_data: List[Dict], converted_data: List[Dict]) -> Dict[str, pd.DataFrame]:
    """，，questionsheet"""
    comparison_data = {}
    
    # question
    original_by_group = {}
    converted_by_group = {}
    
    # Processing
    for item in original_data:
        question_type = item.get('question_type', 'unknown')
        main_type = get_main_question_type(question_type)
        if main_type not in original_by_group:
            original_by_group[main_type] = []
        original_by_group[main_type].append(item)
    
    # Processing
    for item in converted_data:
        question_type = item.get('question_type', 'unknown')
        main_type = get_main_question_type(question_type)
        if main_type not in converted_by_group:
            converted_by_group[main_type] = []
        converted_by_group[main_type].append(item)
    
    # question
    all_groups = set(original_by_group.keys()) | set(converted_by_group.keys())
    
    # question
    for main_type in sorted(all_groups):
        original_items = original_by_group.get(main_type, [])
        converted_items = converted_by_group.get(main_type, [])
        
        # 
        scene_frame_groups = {}
        
        # Processing
        for item in original_items:
            scene_token = item.get('scene_token', '')
            frame_token = item.get('frame_token', '')
            key = f"{scene_token}_{frame_token}"
            if key not in scene_frame_groups:
                scene_frame_groups[key] = {
                    'scene_token': scene_token,
                    'frame_token': frame_token,
                    'original': [],
                    'converted': []
                }
            scene_frame_groups[key]['original'].append(item)
        
        # Processing
        for item in converted_items:
            scene_token = item.get('scene_token', '')
            frame_token = item.get('frame_token', '')
            key = f"{scene_token}_{frame_token}"
            if key not in scene_frame_groups:
                scene_frame_groups[key] = {
                    'scene_token': scene_token,
                    'frame_token': frame_token,
                    'original': [],
                    'converted': []
                }
            scene_frame_groups[key]['converted'].append(item)
        
        # 
        final_rows = []
        
        for key, group in scene_frame_groups.items():
            scene_token = group['scene_token']
            frame_token = group['frame_token']
            original_items = group['original']
            converted_items = group['converted']
            
            # 
            for item in original_items:
                # Processingimage
                object_info = item.get('object_info', {})
                image_paths = item.get('image_path', {})
                
                # ，；，
                if isinstance(object_info, dict):
                    object_info_str = ', '.join([f"{k}: {v}" for k, v in object_info.items()]) if object_info else ""
                else:
                    # Processing
                    if object_info == "N/A" or object_info is None:
                        object_info_str = ""
                    else:
                        object_info_str = str(object_info)
                
                if isinstance(image_paths, dict):
                    image_paths_str = ', '.join([f"{k}: {v}" for k, v in image_paths.items()]) if image_paths else "image"
                else:
                    # Processingimage
                    if image_paths == "N/A" or image_paths is None:
                        image_paths_str = "image"
                    else:
                        image_paths_str = str(image_paths)
                
                row = {
                    'Token': scene_token,
                    'Token': frame_token,
                    '': '',
                    'question': item.get('question_type', ''),
                    'question': item.get('question', ''),
                    '': item.get('answer', ''),
                    '': item.get('correct_answer', ''),
                    '': object_info_str,
                    'image': image_paths_str
                }
                final_rows.append(row)
            
            # ，question
            converted_items.sort(key=lambda x: x.get('question_type', ''))
            for item in converted_items:
                # Processingimage
                object_info = item.get('object_info', {})
                image_paths = item.get('image_path', {})
                
                # ，；，
                if isinstance(object_info, dict):
                    object_info_str = ', '.join([f"{k}: {v}" for k, v in object_info.items()]) if object_info else ""
                else:
                    # Processing
                    if object_info == "N/A" or object_info is None:
                        object_info_str = ""
                    else:
                        object_info_str = str(object_info)
                
                if isinstance(image_paths, dict):
                    image_paths_str = ', '.join([f"{k}: {v}" for k, v in image_paths.items()]) if image_paths else "image"
                else:
                    # Processingimage
                    if image_paths == "N/A" or image_paths is None:
                        image_paths_str = "image"
                    else:
                        image_paths_str = str(image_paths)
                
                row = {
                    'Token': scene_token,
                    'Token': frame_token,
                    '': '',
                    'question': item.get('question_type', ''),
                    'question': item.get('question', ''),
                    '': item.get('answer', ''),
                    '': item.get('correct_answer', ''),
                    '': object_info_str,
                    'image': image_paths_str
                }
                final_rows.append(row)
        
        if final_rows:
            comparison_data[main_type] = pd.DataFrame(final_rows)
    
    return comparison_data

def create_summary_statistics(original_data: List[Dict], converted_data: List[Dict]) -> pd.DataFrame:
    """"""
    # 
    original_stats = {}
    for item in original_data:
        question_type = item.get('question_type', 'unknown')
        original_stats[question_type] = original_stats.get(question_type, 0) + 1
    
    # 
    converted_stats = {}
    for item in converted_data:
        question_type = item.get('question_type', 'unknown')
        converted_stats[question_type] = converted_stats.get(question_type, 0) + 1
    
    # 
    all_types = set(original_stats.keys()) | set(converted_stats.keys())
    summary_rows = []
    
    for question_type in sorted(all_types):
        original_count = original_stats.get(question_type, 0)
        converted_count = converted_stats.get(question_type, 0)
        change = converted_count - original_count
        
        if original_count > 0:
            change_percent = (change / original_count) * 100
        else:
            change_percent = float('inf') if change > 0 else 0
        
        summary_rows.append({
            'question': question_type,
            '': original_count,
            '': converted_count,
            '': change,
            '': f"{change_percent:.1f}%" if change_percent != float('inf') else "N/A"
        })
    
    return pd.DataFrame(summary_rows)

def create_grouped_summary_statistics(original_data: List[Dict], converted_data: List[Dict]) -> pd.DataFrame:
    """"""
    # 
    original_stats = {}
    for item in original_data:
        question_type = item.get('question_type', 'unknown')
        main_type = get_main_question_type(question_type)
        original_stats[main_type] = original_stats.get(main_type, 0) + 1
    
    # 
    converted_stats = {}
    for item in converted_data:
        question_type = item.get('question_type', 'unknown')
        main_type = get_main_question_type(question_type)
        converted_stats[main_type] = converted_stats.get(main_type, 0) + 1
    
    # 
    all_groups = set(original_stats.keys()) | set(converted_stats.keys())
    summary_rows = []
    
    for main_type in sorted(all_groups):
        original_count = original_stats.get(main_type, 0)
        converted_count = converted_stats.get(main_type, 0)
        change = converted_count - original_count
        
        if original_count > 0:
            change_percent = (change / original_count) * 100
        else:
            change_percent = float('inf') if change > 0 else 0
        
        summary_rows.append({
            'question': main_type,
            '': original_count,
            '': converted_count,
            '': change,
            '': f"{change_percent:.1f}%" if change_percent != float('inf') else "N/A"
        })
    
    return pd.DataFrame(summary_rows)

def save_to_excel(comparison_data: Dict[str, pd.DataFrame], 
                 detailed_summary: pd.DataFrame, 
                 grouped_summary: pd.DataFrame, 
                 output_file: str):
    """toExcel"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 
        grouped_summary.to_excel(writer, sheet_name='', index=False)
        
        # 
        detailed_summary.to_excel(writer, sheet_name='', index=False)
        
        # question
        for main_type, df in comparison_data.items():
            # sheet
            sheet_name = f"{main_type}_"
            if len(sheet_name) > 31:  # Excel sheet
                sheet_name = sheet_name[:31]
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Excel：
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    
    wb = load_workbook(output_file)
    
    # 
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 
        column_widths = {
            'A': 20,  # Token
            'B': 20,  # Token
            'C': 10,  # 
            'D': 15,  # question
            'E': 50,  # question
            'F': 30,  # 
            'G': 30,  # 
            'H': 40,  # 
            'I': 40   # image
        }
        
        # 
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 
    wb.save(output_file)
    
    print(f"Excel: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='question - ')
    parser.add_argument('--original_dir', default=DEFAULT_ORIGINAL_DIR, 
                       help=f' (: {DEFAULT_ORIGINAL_DIR})')
    parser.add_argument('--converted_dir', default=DEFAULT_CONVERTED_DIR, 
                       help=f'Converted data directory (: {DEFAULT_CONVERTED_DIR})')
    parser.add_argument('--output_file', default=DEFAULT_OUTPUT_FILE, 
                       help=f'Excel (: {DEFAULT_OUTPUT_FILE})')
    parser.add_argument('--data_variant', default='backlight', 
                       help='variant (: backlight)')
    
    args = parser.parse_args()
    
    # （Project root）
    original_path = os.path.join(args.original_dir, args.data_variant)
    converted_path = os.path.join(args.converted_dir, args.data_variant)
    
    print("=" * 60)
    print("question - ")
    print("=" * 60)
    print(f": {original_path}")
    print(f"Converted data directory: {converted_path}")
    print(f": {args.output_file}")
    print()
    
    # 
    if not os.path.exists(original_path):
        print(f": directory not found: {original_path}")
        return
    
    if not os.path.exists(converted_path):
        print(f": directory not found: {converted_path}")
        return
    
    # 
    print("...")
    original_data = load_json_files(original_path)
    print(f"Done， {len(original_data)} ")
    
    print("...")
    converted_data = load_json_files(converted_path)
    print(f"Done， {len(converted_data)} ")
    
    # 
    print("...")
    comparison_data = create_final_comparison_data(original_data, converted_data)
    
    # 
    print("...")
    detailed_summary = create_summary_statistics(original_data, converted_data)
    grouped_summary = create_grouped_summary_statistics(original_data, converted_data)
    
    # toExcel
    print("Excel...")
    save_to_excel(comparison_data, detailed_summary, grouped_summary, args.output_file)
    
    print()
    print("=" * 60)
    print("Done！")
    print(f"Excel: {args.output_file}")
    print(f" {len(comparison_data)} question")
    print("=" * 60)

if __name__ == "__main__":
    main()
